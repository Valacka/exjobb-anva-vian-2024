import torch
from transformers import AutoTokenizer
from optimum.neuron import NeuronModelForCausalLM
import time
import os
import datetime
import json

import openpyxl

filename = ""
model_name = ""
model = None
name = ""

user_input = ""

# Get user input
while user_input != "1" and user_input != "2":
    user_input = input("Enter 1 for GPT-SW3 20B Instruct, or 2 for Viking 33B: ")

# Process the input using an if statement
if user_input == "1":
    name = "GPT-SW3-"
    model_name = "AI-Sweden-Models/gpt-sw3-20b-instruct"
    model = NeuronModelForCausalLM.from_pretrained("sw3_neuronx_1024")
elif user_input == "2":
    name = "VIKING-"
    model_name = "LumiOpen/Viking-33B"
    model = NeuronModelForCausalLM.from_pretrained("viking_neuronx_1024")

# Function to write data to the file
def write_to_file(data):
    # Open the file in append mode
    with open(filename, 'a') as file:
        file.write(data + '\n\n')

# Initialize Variables
device = "cpu"

# Initialize Tokenizer & Model
tokenizer = AutoTokenizer.from_pretrained(model_name)
tokenizer.pad_token_id = tokenizer.eos_token_id

while True:
    current_time = datetime.datetime.now()
    filename = name+current_time.strftime("%Y-%m-%d_%H-%M-%S.xlsx")

    # Prepare data storage for each type
    data_spelling = []
    data_grammar = []
    data_punctuation = []
    data_misc = []

    json_filename = input("Enter the name of the JSON file to import (or type 'exit' to finish): ")
    if json_filename.lower() == 'exit':
        break

    tasks = []
    try:
        with open(json_filename, 'r') as file:
            data = json.load(file)
    except FileNotFoundError:
        print(f"Error: File {json_filename} not found.")
        continue

    meta = data[0]
    tasks = data[1:]

    iterator = 0
    now = datetime.datetime.now()
    print("="*88)
    current_time = now.strftime("%Y-%m-%d %H:%M:%S")
    print("\nTEST "+model_name+" " * 51 , current_time)

    for task in tasks:
        iterator+=1
        # Process the prompt
        original = task["prompt"]
        prompt = meta["pre_prompt"] + original + meta["post_prompt"]
        
        input_ids = tokenizer(prompt, return_tensors="pt")["input_ids"].to(device)

        start_time = time.time()  # Start timing

        # Generate text
        generated_token_ids = model.generate(
            input_ids,
            max_new_tokens=task["token_length"],
            do_sample=True,
            temperature=0.1,
            top_p=0.95,
            top_k=50,
        )[0]

        end_time = time.time()  # Stop timing after generation
        time_taken = end_time - start_time


        # Decode the generated token IDs back to a string, excluding the prompt
        generated_text = tokenizer.decode(generated_token_ids)
        prompt_length = len(tokenizer.encode(prompt))
        generated_text_without_prompt = tokenizer.decode(generated_token_ids[prompt_length:], skip_special_tokens=True)

        # Print the original and persona-adjusted text and timing
        print("")
        print("-" * 40 +"[TASK "+str(iterator)+"]"+"-" * 40)  # Separator for readability
        print(task["type"])
        print("")
        print("Original Text:", original)
        print("Corrected Text:", generated_text_without_prompt)
        print("")
        print(f"Time taken: {time_taken} seconds")
        print("")  # Separator for readability

        text_type = task["type"]

        if text_type == "Spelling":
            data_spelling.append(generated_text_without_prompt)
        elif text_type == "Grammar":
            data_grammar.append(generated_text_without_prompt)
        elif text_type == "Punctuation":
            data_punctuation.append(generated_text_without_prompt)
        else:
            data_misc.append(generated_text_without_prompt)
            
    # Create a new Excel file and add a sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    # Set column titles
    ws['A1'] = 'Spelling'
    ws['B1'] = 'Grammar'
    ws['C1'] = 'Punctuation'

    # Write data to each column
    for idx, value in enumerate(data_spelling, start=2):
        ws[f'A{idx}'] = value

    for idx, value in enumerate(data_grammar, start=2):
        ws[f'B{idx}'] = value

    for idx, value in enumerate(data_punctuation, start=2):
        ws[f'C{idx}'] = value
    for idx, value in enumerate(data_misc, start=2):
        ws[f'D{idx}'] = value

    # Save the workbook
    wb.save(filename)

print("="*88)